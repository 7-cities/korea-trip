"""Sync 2026-06-08 sheet update into places.json.

Naver List had a category-normalization pass: ~30 distinct (old type -> new type)
renames affecting ~67 rows, plus a few EN-name parenthetical clarifiers.

This script:
  1. Builds the (kr_name -> new_type, new_en) map directly from the fresh sheet
  2. For every place in places.json whose name_kr appears in that map, updates
     `type` (and `name_en` if a parenthetical clarifier was added)
  3. Backs up places.json -> places.prev.json
  4. Rewrites places.json + ../web/places_data.js
"""
import json
import shutil
from pathlib import Path
from openpyxl import load_workbook

HERE = Path(__file__).parent
ROOT = HERE.parent
WEB = ROOT / "web"

places = json.loads((ROOT / "places.json").read_text(encoding="utf-8"))

# Build authoritative map from the latest sheet: name_kr -> (type, name_en)
# Tab structures differ:
#   Naver List : col0=Type      col2=name_kr col3=name_en
#   Jeju Island: col0=Location  col1=Type    col2=name_kr col3=name_en
wb = load_workbook(HERE / "sheet.xlsx", data_only=True)
authoritative = {}
TAB_LAYOUTS = {
    "Naver List":  {"type": 0, "kr": 2, "en": 3},
    "Jeju Island": {"type": 1, "kr": 2, "en": 3},
}
for sname, cols in TAB_LAYOUTS.items():
    if sname not in wb.sheetnames:
        continue
    ws = wb[sname]
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 4:
            continue
        def cell(i):
            return (row[i] or "").strip() if isinstance(row[i], str) else ""
        ptype = cell(cols["type"])
        kr = cell(cols["kr"])
        en = cell(cols["en"])
        if not kr or not any('가' <= c <= '힣' for c in kr):
            continue
        if ptype.lower() in {"type", "place", "places", "location"}:
            continue
        authoritative[kr] = {"type": ptype, "name_en": en}

print(f"Authoritative entries from sheet: {len(authoritative)}")

# Apply to places.json
type_changes = 0
name_changes = 0
unmatched = []
for p in places:
    kr = p.get("name_kr") or ""
    if not kr:
        continue
    auth = authoritative.get(kr.strip())
    if not auth:
        # Try a relaxed match (handle trailing spaces)
        for k, v in authoritative.items():
            if k.strip() == kr.strip():
                auth = v
                break
    if not auth:
        unmatched.append(kr)
        continue

    new_type = auth["type"]
    # Normalize the sheet typo
    if new_type == "Restauarnt":
        new_type = "Restaurant"

    if new_type and new_type != (p.get("type") or ""):
        old = p.get("type")
        p["type"] = new_type
        type_changes += 1
        print(f"  type: {kr:30s}  {old!r:30s} -> {new_type!r}")

    # Adopt name_en update only when sheet adds info in parens (e.g. "(bookstore)")
    new_en = auth["name_en"]
    cur_en = (p.get("name_en") or "").strip()
    if new_en and "(" in new_en and new_en.strip().rstrip(")") != cur_en.rstrip(")"):
        if cur_en and new_en.lower().startswith(cur_en.lower().split("(")[0].rstrip().lower()):
            p["name_en"] = new_en
            name_changes += 1
            print(f"  name: {kr:30s}  {cur_en!r} -> {new_en!r}")

print()
print(f"Type changes:  {type_changes}")
print(f"Name changes:  {name_changes}")
print(f"places not in sheet (kept as-is): {len(unmatched)}")

# Back up + write
shutil.copy2(ROOT / "places.json", HERE / "places.prev.json")
(ROOT / "places.json").write_text(json.dumps(places, ensure_ascii=False, indent=2), encoding="utf-8")
js = "const PLACES = " + json.dumps(places, ensure_ascii=False, indent=2) + ";\n"
(WEB / "places_data.js").write_text(js, encoding="utf-8")
print(f"\nWrote places.json ({len(places)} entries) + web/places_data.js")

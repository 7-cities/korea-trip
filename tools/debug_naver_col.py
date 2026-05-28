"""Investigate Column F of 'Naver List' to find where the hyperlinks live."""
from pathlib import Path
from openpyxl import load_workbook

HERE = Path(__file__).parent

# Load with formulas visible (not evaluated values)
wb_f = load_workbook(HERE / "sheet.xlsx", data_only=False)
ws = wb_f["Naver List"]

print(f"Naver List: rows={ws.max_row}, cols={ws.max_column}")
print(f"Inspecting columns A..G for first 10 data rows.\n")

for r in range(2, 12):
    print(f"--- Row {r} ---")
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        formula_or_val = cell.value
        hl = cell.hyperlink.target if cell.hyperlink else None
        print(f"  col{c} (={chr(64+c)}): val={formula_or_val!r}  hyperlink={hl!r}")
    print()

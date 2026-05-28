"""
Extract per-cell URLs from sheet.xlsx, capturing BOTH:
  - Cells with explicit Excel hyperlinks (cell.hyperlink.target)
  - Cells with =HYPERLINK("url", "label") formulas (need data_only=False)

Output: sheet_hyperlinks.json — list of rows with both raw values and links.
"""
import json
import re
from pathlib import Path
from openpyxl import load_workbook

HERE = Path(__file__).parent

# Two passes: data_only=True for evaluated values, False for formulas
wb_eval = load_workbook(HERE / "sheet.xlsx", data_only=True)
wb_form = load_workbook(HERE / "sheet.xlsx", data_only=False)

HYPERLINK_RE = re.compile(r'=HYPERLINK\("([^"]+)"\s*,\s*"([^"]*)"\)', re.IGNORECASE)

out = []
for sname in wb_eval.sheetnames:
    ws_e = wb_eval[sname]
    ws_f = wb_form[sname]
    print(f"\n=== {sname}: {ws_e.max_row} rows x {ws_e.max_column} cols ===")
    link_count = 0
    for r in range(1, ws_e.max_row + 1):
        values = []
        links = {}
        for c in range(1, ws_e.max_column + 1):
            cell_e = ws_e.cell(row=r, column=c)
            cell_f = ws_f.cell(row=r, column=c)
            values.append(cell_e.value)
            # Source 1: explicit cell hyperlink
            if cell_e.hyperlink and getattr(cell_e.hyperlink, "target", None):
                links[c - 1] = cell_e.hyperlink.target
                continue
            # Source 2: HYPERLINK() formula in the unevaluated workbook
            v = cell_f.value
            if isinstance(v, str):
                m = HYPERLINK_RE.search(v)
                if m:
                    links[c - 1] = m.group(1)
        if any(v for v in values) or links:
            out.append({
                "sheet": sname,
                "row": r,
                "values": values,
                "links": links,
            })
        link_count += len(links)
    print(f"  hyperlinks: {link_count}")

(HERE / "sheet_hyperlinks.json").write_text(
    json.dumps(out, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
)
total_links = sum(len(r["links"]) for r in out)
print(f"\nTotal rows with content: {len(out)}, total hyperlinks: {total_links}")
